#CREACIÓN DE UNA BASE DE DATOS CON DOS TABLAS RELACIONADAS UNO A MUCHOS
import random as rd
import sys
import datetime
import sqlite3
from sqlite3 import Error
from datetime import (date, 
                      datetime,
                      timezone,
                      timedelta)
import openpyxl

#Crea una tabla en SQLite3
def Crear_tabla ():
  try:
    with sqlite3.connect("Grupo34.db") as conn:
          mi_cursor = conn.cursor()
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS Usuarios (clave INTEGER PRIMARY KEY, nombre TEXT NOT NULL);")
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS Salas (clave INTEGER PRIMARY KEY, nombre TEXT NOT NULL, capacidad INTEGER NOT NULL);")
          mi_cursor.execute("CREATE TABLE IF NOT EXISTS Reservaciones (folio INTEGER PRIMARY KEY, nombre TEXT NOT NULL, horario Text NOT NULL, fecha timestamp) ")
          print("Tablas creadas exitosamente")
  except Error as e:
      print (e)
  except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
  finally:
      conn.close()

Crear_tabla ()

tablas_a_exportar = ['Reservaciones', 'Salas', 'Usuarios'] 
nombre_de_archivo = 'Grupo34.xlsx'

# Opcion 1
def Registrar_Reservacion ():
    while True:
        valor_clave = int(input("Ingresa la clave de cliente: "))
        try:
            with sqlite3.connect("Grupo34.db") as conn:
                mi_cursor = conn.cursor()
                valores = {"clave":valor_clave}
                mi_cursor.execute("SELECT * FROM Usuarios WHERE clave = :clave", valores)
                registro = mi_cursor.fetchall()

                if registro:
                    for clave, nombre in registro:
                        print(f"{clave}\t{nombre}")
                else:
                    print(f"No se encontró un proyecto asociado con la clave de cliente {valor_clave}")
        except Error as e:
            print (e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        else:
            Nombre = input("Ingresa el nombre de la reservación (Escribe SALIR para regresar al menú): ")
            if Nombre == '':
                continue
            elif Nombre.upper() == 'SALIR':
                mostrar_menu()
                return
            else:
                Horario = input("Ingresa el horario deseado [M, V, N]: ")
                print(Horario)
                Fecha_Ingresada= input("Ingresa la fecha de reservación: ")
                Fecha_dt = datetime.strptime(Fecha_Ingresada, '%d/%m/%Y')
                fecha_actual = datetime.today()
                fecha_permitida = datetime.now() + timedelta( days = 2)
                if Fecha_dt < fecha_permitida:
                    print("Debes hacer la reservacion con 2 dias de anticipación.")
                else:
                    nivel = rd.randint(1,99)
                    try:
                        with sqlite3.connect("Grupo34.db") as conn:
                            mi_cursor = conn.cursor()
                            Grupo34={"folio":nivel,"nombre":Nombre,"horario":Horario,"fecha":Fecha_dt}
                            mi_cursor.execute("INSERT INTO Reservaciones VALUES(:folio,:nombre,:horario,:fecha)",Grupo34)
                            #print("El folio es ")
                            #print(nivel)
                            #print("El nombre es ")
                            #print(Nombre)
                    except Error as e:
                        print (e)
                    except:
                        print(f"Surgio una falla siendo esta la causa: {sys.exc_info()[0]}")
                    finally:
                        if (conn):
                            conn.close()
                            fecha_consultar = input("Confirma la fecha (dd/mm/aaaa): ")
                            fecha_consultar = datetime.strptime(fecha_consultar, "%d/%m/%Y").date()
                            print("¡Reservacion realizada con exito!")
                            try:
                                with sqlite3.connect("Grupo34.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                                    mi_cursor = conn.cursor()
                                    criterios = {"fecha":fecha_consultar}
                                    mi_cursor.execute("SELECT folio, nombre, horario, fecha FROM Reservaciones WHERE DATE(fecha) = (:fecha);", criterios)
                                    #mi_cursor.execute("SELECT clave, nombre, fecha_registro FROM Amigo WHERE DATE(fecha_registro) >= :fecha;", criterios)
                                    registros = mi_cursor.fetchall()
   
                                    for clave, nombre, fecha_registro, fecha in registros:
                                        print(f"Clave = {clave}, tipo de dato {type(clave)}")
                                        print(f"Nombre = {nombre}, tipo de dato {type(nombre)}")
                                        print(f"Horario = {fecha_registro}, tipo de dato {type(fecha_registro)}")
                                        print(f"Fecha = {fecha}, tipo de dato {type(fecha)}")
                                    for clave, nombre, fecha_registro, fecha in registros:
                                        print("Clave\t" + "Nombre\t"+ " Turno\t" + "            Fecha\t")
                                        print(f"{clave}\t {nombre}\t {fecha_registro}\t {fecha}\t")
                                        return

                            except sqlite3.Error as e:
                                print (e)
                            except Exception:
                                print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                            finally:
                                if (conn):
                                    conn.close()
                                    print("Se ha cerrado la conexión")
                                    print("*" * 20)
                                    print("1. Registra la reservacion\n"
                                    "2. Modificar las descripciones de la reservacion\n" +
                                    "3. Consulta la fecha disponible\n" +
                                    "4. Reporte de la reservaciones de una fecha\n" +
                                    "5. Registrar Sala\n" +
                                    "6. Registrar Cliente\n" +
                                    "7. Salir del programa\n" +
                                    "8. Eliminar reservacion \n" +
                                    "9. Exportar base de datos a Excel \n")

#Opcion 2
def modificar_descripciones ():
    while True:
        llave = input("Ingresa el nombre de tu sala actual: ")
        try:
            with sqlite3.connect("Grupo34.db") as conn:
                mi_cursor = conn.cursor()
                valores1 = {"nombre":llave}
                mi_cursor.execute("SELECT * FROM Reservaciones WHERE nombre = :nombre", valores1)
                registro = mi_cursor.fetchall()

                if registro:
                    for folio, nombre, horario, fecha in registro:
                        print("Clave\t" + "Nombre\t"+ " Turno\t" + "            Fecha\t")
                        print(f"{folio}\t{nombre}\t{horario}\t{fecha}")
                else:
                    print(f"No se encontró un proyecto asociado con la clave {llave}")
        except Error as e:
            print (e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        else:
            nuevo_nombre = input("Ingrese el nuevo nombre: ")
            id_number = folio
            Turno = horario
            fecha_dt = fecha
            try:
                with sqlite3.connect("Grupo34.db") as conn:
                    mi_cursor = conn.cursor()
                    Sydney = {"folio":id_number, "nombre":nuevo_nombre,"turno":Turno,"fecha":fecha_dt}
                    mi_cursor.execute("UPDATE Reservaciones SET nombre = (:nombre) WHERE (folio) = (:folio);", Sydney)
                    print("Modificacion realizada con exito.")
                    return
            except Error as e:
                print (e)
            except:
                print(f"Surgio una falla siendo esta la causa: {sys.exc_info()[0]}")

#opcion 3

import sqlite3
from datetime import datetime
from sqlite3 import Error

def consulta_fecha():
    while True:
        fecha_consultar = input("Ingrese la fecha a consultar (aaaa-mm-dd): ")
        try:
            fecha_dt = datetime.strptime(fecha_consultar, '%Y-%m-%d')
            break  # Si la fecha es válida, salir del bucle
        except ValueError:
            print("Formato de fecha incorrecto. Por favor, ingrese una fecha válida en el formato dd/mm/aaaa.")

    try:
        with sqlite3.connect("Grupo34.db", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            criterios = {"fecha": fecha_consultar}
            mi_cursor.execute("SELECT * FROM Reservaciones WHERE DATE(fecha) = :fecha",criterios)
            registros = mi_cursor.fetchall()
            if registros:
                print("Estas son las fechas ocupadas:")
                for folio, nombre, horario, fecha in registros:
                    print(f"Folio: {folio}, Nombre: {nombre}, Horario: {horario}, Fecha: {fecha}")
            else:
                print("No se encontraron reservaciones para la fecha especificada.")
    except Error as e:
        print(e)
    except Exception as ex:
        print(f"Se produjo el siguiente error: {ex}")
    finally:
        if conn:
            conn.close()
            print("Se ha cerrado la conexión")      


#Opcion 4
def reporte_a_Excel ():
    while True:
        fecha_reporte = input("¿De que fecha quieres sacar el reporte? (Escribe SALIR si deseas volver al menú): ")
        if fecha_reporte.upper() == "SALIR":
            mostrar_menu()
            return
        fecha_reporte = datetime.strptime(fecha_reporte, "%d/%m/%Y")
        print("*"* 75)
        print("**            REPORTE DE RESERVACIONES PARA EL DIA", fecha_reporte, "           **")
        try:
            with sqlite3.connect("Grupo34.db", detect_types = sqlite3.PARSE_DECLTYPES  | sqlite3.PARSE_COLNAMES) as conn:
                mi_cursor = conn.cursor()
                criterios = {"fecha":fecha_reporte}
                mi_cursor.execute("SELECT folio, nombre, horario, fecha FROM Reservaciones WHERE fecha = :fecha", criterios)
                registrados = mi_cursor.fetchall()
                
                if registrados:
                    lista = []  # Inicializa la lista fuera del bucle
                    for folio, nombre, horario, fecha in registrados:
                        print("Clave\t" + "Nombre\t" + "Turno\t" + "Fecha\t")
                        Campos = ("Clave", "Nombre", "Turno", "Fecha")
                        print(f"{folio}\t{nombre}\t{horario}\t{fecha}")
                        reporte = (folio, nombre, horario, fecha)  # Crear una tupla con los datos
                        lista.append(reporte)
                        print(lista)
                        wb = openpyxl.Workbook()
                        hoja = wb.active
                        for listas in lista:
                            hoja.append(Campos)
                            hoja.append(listas)

                        print(f'Hoja activa: {hoja.title}')
                        hoja["A1"] = 10
                        a1 = hoja["A1"]
                        print(a1.value)
                        wb.save('proyectos.xlsx')
                        print("Reporte exportado en archivo proyectos.xlsx")
                else:
                    print(f"No se encontraron reservaciones para la fecha {fecha_reporte}")
                    return
        except Error as e:
            print (e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        
            
#Opcion 5
def Registrar_Sala ():
    while True:
        SALA = input("Ingresa el nombre de la sala (Escribe SALIR para regresar al menú): ")
        if SALA =='':
            break
        elif SALA.upper() =='SALIR':
            mostrar_menu()
            return
        else:
            capacity = int(input("Ingresa la capacidad de la sala: "))
            N = rd.randint(1,99)
            try:
                with sqlite3.connect("Grupo34.db") as conn:
                    mi_cursor = conn.cursor()
                    Valores5={"clave":N,"nombre":SALA,"capacidad":capacity }
                    mi_cursor.execute("INSERT INTO Salas (clave, nombre, capacidad) VALUES(:clave,:nombre,:capacidad)",Valores5)
                    print("Sala Registrada!!")
                    print("Tu clave de la Sala es la siguiente")
                    print(N)
            except Error as e:
                print (e)
            except:
                print(f"Surgio una falla siendo esta la causa: {sys.exc_info()[0]}")
            finally:
                if (conn):
                    conn.close()

# Opcion 6
def Registrar_Cliente ():
    while True:
        Usuario=input("Ingresa el nombre del usuario (Escribe SALIR si quieres regresar al menú principal): ")
        n = rd.randint(1,99)
        if Usuario == '':
            break
        elif Usuario.upper() == 'SALIR':
            mostrar_menu()
            return
        else:
            try:
                with sqlite3.connect("Grupo34.db") as conn:
                    mi_cursor = conn.cursor()
                    valores={"clave":n,"nombre":Usuario} 
                    mi_cursor.execute("INSERT INTO Usuarios (clave, nombre) VALUES(:clave,:nombre)",valores)
                print("Usuario registrado!")
                print("Tu clave de cliente es la siguiente: ")
                print(n)
            except Error as e:
                print (e)
            except:
                print(f"Surgio una falla siendo esta la causa: {sys.exc_info()[0]}")
            finally:
                if (conn):
                    conn.close()

# Opcion 7
def salir_del_programa ():
    print ("*"*30)
    print ("Hasta pronto tenga un buen dia :D")
    print ("*"*30)
    print ("Vuelve a visitarnos pronto")
    print ("*"*30)
    
# Opcion 8
def eliminar_reservacion ():
    while True:
        key_code = int(input("Ingresa la clave de cliente: "))
        try:
            with sqlite3.connect("Grupo34.db") as conn:
                mi_cursor = conn.cursor()
                valores = {"clave":key_code}
                mi_cursor.execute("SELECT * FROM Usuarios WHERE clave = :clave", valores)
                registro = mi_cursor.fetchall()

                if registro:
                    for clave, nombre in registro:
                        print(f"{clave}\t{nombre}")
                else:
                    print(f"No se encontró un proyecto asociado con la clave {key_code}")
        except Error as e:
            print (e)
        except:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
        else:
            id_sala = input("Ingrese el ID de la reservación a eliminar (Escribe SALIR para regresar al programa): ")
            if id_sala == '':
                continue
            elif id_sala.upper() == 'SALIR':
                mostrar_menu()
                return
            else:
                Name = input("Cual era tu nombre del evento?: ")
                Turno = input("Escribe tu horario: ")
                fecha_asignada = input("Ingresa la fecha por favor: ")
                fecha_dt = datetime.strptime(fecha_asignada, '%d/%m/%Y')
                fecha_permitida = datetime.now() + timedelta( days = 3)
                if fecha_dt < fecha_permitida:
                    print("Lo siento pero no la puedes cancelar, debes cancelarla con 3 dias de anticipación")
                else:
                    try:
                        with sqlite3.connect("Grupo34.db") as conn:
                            mi_cursor = conn.cursor()
                            delete = {"folio":id_sala,"nombre":Name,"horario":Turno,"fecha":fecha_dt}
                            mi_cursor.execute("DELETE FROM Reservaciones WHERE folio = :folio;", delete)
                            #delete={"folio":id_sala,"nombre":Name,"horario":Turno,"fecha":fecha_dt}
                            #mi_cursor.execute("DELETE FROM Reservaciones WHERE (fecha) = (:fecha);", delete)
                            print("Reservacion eliminada. ¡Lamentamos que hayas decidido cancelar tu evento!")
                            return
                    except sqlite3.Error as e:
                        print (e)
                    except Exception:
                        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
                    finally:
                        if (conn):
                            conn.close()
                            print("Se ha cerrado la conexión")
                        

def exportar_tablas_a_excel(nombres_tablas, nombre_archivo):
    # Crear un nuevo libro de trabajo Excel
    wb = openpyxl.Workbook()

    # Conectarse a la base de datos SQLite
    conn = sqlite3.connect('Grupo34.db')  # Reemplaza 'tu_base_de_datos.db' con el nombre de tu archivo de base de datos
    cursor = conn.cursor()

    for tabla in nombres_tablas:
        # Ejecutar una consulta SQL para obtener datos de la tabla actual
        cursor.execute(f'SELECT * FROM {tabla}')
        datos = cursor.fetchall()

        # Crear una nueva hoja de cálculo para la tabla actual
        hoja = wb.create_sheet(title=tabla)

        # Agregar encabezados a la hoja de cálculo
        encabezados = [description[0] for description in cursor.description]
        hoja.append(encabezados)

        # Agregar datos a la hoja de cálculo
        for fila in datos:
            hoja.append(fila)

    # Guardar el libro de trabajo en un archivo
    wb.save(nombre_archivo)

    # Cerrar la conexión a la base de datos
    conn.close()

    print(f"Tablas exportadas a '{nombre_archivo}' correctamente")


def mostrar_menu():
    print("*" * 20)
    print("1. Registra la reservacion\n"
    "2. Modificar las descripciones de la reservacion\n" +
    "3. Consulta la fecha disponible\n" +
    "4. Reporte de la reservaciones de una fecha\n" +
    "5. Registrar Sala\n" +
    "6. Registrar Cliente\n" +
    "7. Salir del programa\n" +
    "8. Eliminar reservacion \n" +
    "9. Exportar base de datos a Excel \n")


def Mi_menú ():
    print("Conexion Establecida")
    #EstablecerConexion ()
    Crear_tabla ()
    print("*" * 20)
    print("1. Registra la reservacion\n"
    "2. Modificar las descripciones de la reservacion\n" +
    "3. Consulta la fecha disponible\n" +
    "4. Reporte de la reservaciones de una fecha\n" +
    "5. Registrar Sala\n" +
    "6. Registrar Cliente\n" +
    "7. Salir del programa\n" +
    "8. Eliminar reservacion \n" +
    "9. Exportar base de datos a Excel \n")
    while True:
        try:
            Opcion = int(input("Seleccione el numero de la accion que quiere realizar \n:"))
        except Error as e:
            print(e)
        except:
            print(f"Ocurrió un problema {sys.exc_info()[0]}")
        else:
            if Opcion == 1:
                Registrar_Reservacion ()
            elif Opcion == 2:
                modificar_descripciones ()
            elif Opcion == 3:
                consulta_fecha()
            elif Opcion == 4:
                reporte_a_Excel ()
            elif Opcion == 5:
                Registrar_Sala ()
            elif Opcion == 6:
                Registrar_Cliente ()
            elif Opcion == 7:
                salir_del_programa ()
                break
            elif Opcion == 8:
                eliminar_reservacion ()
            elif Opcion == 9:
                exportar_tablas_a_excel(tablas_a_exportar, nombre_de_archivo)
            else:
                print("Eso no esta disponible checa el menú")
    

Mi_menú ()
